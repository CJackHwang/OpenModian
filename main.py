#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
桌游市场调研工具 - 融合版爬虫
整合了原始版本的完整数据提取功能和增强版的优秀架构设计

主要功能：
1. 完整的摩点众筹数据爬取（34个数据字段）
2. 详细的项目状态分析和作者信息提取
3. 多格式输出支持（Excel、JSON、CSV）
4. 完善的错误处理和日志记录
5. 配置化管理和统计监控
6. 现代化的代码架构和网络请求处理
"""

import random
import re
import socket
import time
import ssl
import json
import csv
import logging
import datetime
import urllib.request
import urllib.error
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import requests
import xlwt
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('modian_spider.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ModianSpiderConfig:
    """摩点爬虫配置类"""
    
    def __init__(self, category: str = "all"):
        # 基础URL配置
        self.BASE_DOMAIN = "https://zhongchou.modian.com"
        self.AUTHOR_API_URL = "https://apim.modian.com/apis/comm/user/user_info"

        # 分类URL映射（基于摩点网站实际分类）
        self.CATEGORY_URLS = {
            # 基础分类
            "all": "/all/top_time/all/",
            "success": "/all/top_time/success/",
            "going": "/all/top_time/going/",
            "preheat": "/all/top_time/preheat/",
            "idea": "/all/top_time/idea/",

            # 具体项目分类
            "games": "/games/top_time/all/",
            "publishing": "/publishing/top_time/all/",
            "tablegames": "/tablegames/top_time/all/",
            "toys": "/toys/top_time/all/",
            "cards": "/cards/top_time/all/",
            "technology": "/technology/top_time/all/",
            "film-video": "/film-video/top_time/all/",
            "music": "/music/top_time/all/",
            "activities": "/activities/top_time/all/",
            "design": "/design/top_time/all/",
            "curio": "/curio/top_time/all/",
            "home": "/home/top_time/all/",
            "food": "/food/top_time/all/",
            "comics": "/comics/top_time/all/",
            "charity": "/charity/top_time/all/",
            "animals": "/animals/top_time/all/",
            "wishes": "/wishes/top_time/all/",
            "others": "/others/top_time/all/"
        }

        # 设置当前分类
        self.category = category if category in self.CATEGORY_URLS else "all"
        self.BASE_URL = f"{self.BASE_DOMAIN}{self.CATEGORY_URLS[self.category]}"
        
        # 输出配置
        self.OUTPUT_DIR = Path("output")
        self.CACHE_DIR = Path("cache")
        self.EXCEL_FILENAME = "摩点众筹-主要信息.xls"
        
        # 请求配置
        self.MAX_RETRIES = 5
        self.RETRY_DELAY = 2
        self.REQUEST_TIMEOUT = (10, 20)
        self.MAX_PAGES = 3  # 默认测试范围，可修改为更大值如833
        self.SAVE_INTERVAL = 5  # 每5页保存一次
        
        # 请求头配置
        self.HEADERS = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1"
        }
        
        # 移动端请求头（用于作者页面）
        self.MOBILE_HEADERS = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }
        
        # API请求头
        self.API_HEADERS = {
            'Accept': 'application/json, text/plain, */*',
            'User-Agent': 'Mozilla/5.0 (Linux; Android 10; SM-G975F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Mobile Safari/537.36',
            'Origin': 'https://m.modian.com',
            'Referer': 'https://m.modian.com/',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        }
        
        # 正则表达式
        self.USER_ID_PATTERN = re.compile(r'https://me.modian.com/u/detail\?uid=(\d+)')
        self.LINK_ID_PATTERN = re.compile(r'https://zhongchou.modian.com/item/(\d+).html')
        
        # 创建必要目录
        self.OUTPUT_DIR.mkdir(exist_ok=True)
        self.CACHE_DIR.mkdir(exist_ok=True)

    def get_page_url(self, page: int) -> str:
        """获取指定页面的URL"""
        return f"{self.BASE_URL}{page}"

    def set_category(self, category: str):
        """设置爬取分类"""
        if category in self.CATEGORY_URLS:
            self.category = category
            self.BASE_URL = f"{self.BASE_DOMAIN}{self.CATEGORY_URLS[category]}"
        else:
            logger.warning(f"未知分类: {category}，使用默认分类 'all'")
            self.category = "all"
            self.BASE_URL = f"{self.BASE_DOMAIN}{self.CATEGORY_URLS['all']}"


class ModianSpiderStats:
    """爬虫统计类"""
    
    def __init__(self):
        self.start_time = time.time()
        self.total_requests = 0
        self.successful_requests = 0
        self.failed_requests = 0
        self.projects_found = 0
        self.projects_processed = 0
        self.pages_processed = 0
        self.errors = {}
    
    def record_request(self, success: bool):
        """记录请求结果"""
        self.total_requests += 1
        if success:
            self.successful_requests += 1
        else:
            self.failed_requests += 1
    
    def record_error(self, error_type: str, message: str):
        """记录错误"""
        if error_type not in self.errors:
            self.errors[error_type] = []
        self.errors[error_type].append(message)
    
    def get_summary(self) -> Dict[str, Any]:
        """获取统计摘要"""
        elapsed_time = time.time() - self.start_time
        return {
            "elapsed_time": elapsed_time,
            "total_requests": self.total_requests,
            "successful_requests": self.successful_requests,
            "failed_requests": self.failed_requests,
            "success_rate": (self.successful_requests / max(self.total_requests, 1)) * 100,
            "projects_found": self.projects_found,
            "projects_processed": self.projects_processed,
            "pages_processed": self.pages_processed,
            "avg_time_per_page": elapsed_time / max(self.pages_processed, 1),
            "errors": self.errors
        }


class ModianSpider:
    """摩点众筹爬虫主类"""
    
    def __init__(self, config: Optional[ModianSpiderConfig] = None):
        self.config = config or ModianSpiderConfig()
        self.stats = ModianSpiderStats()
        self.session = requests.Session()
        self.session.headers.update(self.config.HEADERS)
        
        # SSL配置
        self.ssl_context = ssl.create_default_context()
        self.ssl_context.check_hostname = False
        self.ssl_context.verify_mode = ssl.CERT_NONE
        
        logger.info("摩点众筹爬虫初始化完成")
    
    def make_request(self, url: str, use_mobile: bool = False) -> Optional[str]:
        """发起网络请求"""
        headers = self.config.MOBILE_HEADERS if use_mobile else self.config.HEADERS
        
        for attempt in range(self.config.MAX_RETRIES):
            try:
                timeout = random.randint(*self.config.REQUEST_TIMEOUT)
                
                # 使用requests进行请求
                response = self.session.get(url, headers=headers, timeout=timeout, verify=False)
                response.raise_for_status()
                response.encoding = 'utf-8'
                
                self.stats.record_request(True)
                return response.text
                
            except Exception as e:
                logger.warning(f"第{attempt + 1}次请求失败 {url}: {e}")
                self.stats.record_request(False)
                self.stats.record_error("network_error", str(e))
                
                if attempt == self.config.MAX_RETRIES - 1:
                    logger.error(f"请求最终失败 {url}")
                    return None
                
                wait_time = self.config.RETRY_DELAY * (attempt + 1)
                logger.info(f"等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)
        
        return None
    
    def askURL(self, url):
        """兼容旧版本的请求方法（使用urllib）"""
        head = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        timeout_range = (10, 20)

        html = ""
        for i in range(5):  # Increased retry attempts
            try:
                timeout = random.randint(*timeout_range)
                request = urllib.request.Request(url, headers=head)
                response = urllib.request.urlopen(request, timeout=timeout, context=self.ssl_context)
                html = response.read().decode("utf-8")
                self.stats.record_request(True)
                break
            except (urllib.error.URLError, ConnectionResetError, socket.timeout, ssl.SSLError) as e:
                logger.warning(f'第{i + 1}次尝试失败，原因：{e} URL: {url}')
                self.stats.record_request(False)
                self.stats.record_error("network_error", str(e))
                if i == 4:  # Last attempt
                    logger.error(f'重试多次仍然失败！URL: {url}')
                    break
                # Exponential backoff
                wait_time = (i + 1) * 2
                logger.info(f'等待 {wait_time} 秒后重试...')
                time.sleep(wait_time)
        return html

    def askURL2(self, url):
        """移动端请求方法（用于作者页面）"""
        head = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1'
        }
        timeout_range = (10, 20)

        html = ""
        for i in range(5):  # Increased retry attempts
            try:
                timeout = random.randint(*timeout_range)
                request = urllib.request.Request(url, headers=head)
                response = urllib.request.urlopen(request, timeout=timeout, context=self.ssl_context)
                html = response.read().decode("utf-8")
                self.stats.record_request(True)
                break
            except (urllib.error.URLError, ConnectionResetError, socket.timeout, ssl.SSLError) as e:
                logger.warning(f'第{i + 1}次尝试(askURL2)失败，原因：{e} URL: {url}')
                self.stats.record_request(False)
                self.stats.record_error("network_error", str(e))
                if i == 4:  # Last attempt
                    logger.error(f'重试多次(askURL2)仍然失败！URL: {url}')
                    break
                # Exponential backoff
                wait_time = (i + 1) * 2
                logger.info(f'等待 {wait_time} 秒后重试...')
                time.sleep(wait_time)
        return html

    def get_author_info_from_api(self, uid):
        """从API获取作者信息"""
        params = {"json_type": 1, "to_user_id": uid, "user_id": uid}
        headers = self.config.API_HEADERS.copy()
        headers['Timestamp'] = str(int(time.time()))

        try:
            response = requests.get(self.config.AUTHOR_API_URL, params=params, headers=headers, timeout=10)
            response.raise_for_status()
            self.stats.record_request(True)
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching author API for UID {uid}: {e}")
            self.stats.record_request(False)
            self.stats.record_error("api_error", str(e))
            return {}

    def parse_author_page_info(self, html, author_id_from_url):
        """解析作者页面信息"""
        soup = BeautifulSoup(html, "html.parser")
        author_data = []

        fans_num = 0
        notice_number = 0
        love_number = 0

        banner_div = soup.find('div', {'class': 'banner'})
        if banner_div:
            cont_div = banner_div.find('div', {'class': 'cont'})
            if cont_div:
                fans_span = cont_div.find('span', {'class': 'go_span fans'})
                if fans_span and fans_span.find('i'):
                    fans_num_text = fans_span.find('i').text.strip()
                    if fans_num_text.isdigit():
                        fans_num = int(fans_num_text)

                notice_span = cont_div.select_one('span.go_span:not(.fans)')
                if notice_span:
                    text = notice_span.text.strip()
                    parts = text.split()
                    if parts and parts[0].isdigit():
                        notice_number = int(parts[0])

                all_span = cont_div.find('span', {'id': 'ALL'})
                if all_span:
                    text2 = all_span.text.strip()
                    parts2 = text2.split()
                    if parts2 and parts2[0].isdigit():
                        love_number = int(parts2[0])

        author_data.extend([fans_num, notice_number, love_number])

        detail_result = {}
        detail_div = soup.find('div', {'class': 'detail'})
        if detail_div:
            for item_div in detail_div.find_all('div', class_='item'):
                label_tag = item_div.find('label')
                p_tag = item_div.find('p')
                if label_tag and p_tag:
                    detail_result[label_tag.text.strip()] = p_tag.text.strip()
        author_data.append(str(detail_result))

        other_result = {}
        other_info_div = soup.find('div', {'class': 'other_info'})
        if other_info_div:
            for item_div in other_info_div.find_all('div', class_='item'):
                p_tags = item_div.find_all('p')
                if len(p_tags) == 2 and p_tags[1].text.strip().isdigit():
                    value = int(p_tags[1].text.strip())
                    key = p_tags[0].text.strip()
                    other_result[key] = value
        author_data.append(str(other_result))

        userhome_url = f"https://m.modian.com/user/homePage/{author_id_from_url}"
        author_data.append(userhome_url)

        return author_data

    def get_project_status_info(self, soup_detail_page):
        """获取项目状态信息"""
        status_info = {
            "item_class": "未知情况",
            "is_idea": False,
            "is_preheat": False,
            "is_going": False,
            "is_success": False,
            "is_fail": False # Added for clarity
        }
        buttons_div = soup_detail_page.find('div', {'class': 'buttons clearfloat'})
        if buttons_div:
            button_a = buttons_div.select_one('a')
            if button_a:
                class_result_text = button_a.text.strip()
                status_info["raw_status_text"] = class_result_text # Store raw text for debugging
                if class_result_text == "看好":
                    status_info["item_class"] = "创意"
                    status_info["is_idea"] = True
                elif class_result_text == "看好项目":
                    status_info["item_class"] = "预热"
                    status_info["is_preheat"] = True
                elif class_result_text == "立即购买支持":
                    status_info["item_class"] = "众筹中"
                    status_info["is_going"] = True
                elif class_result_text == "众筹成功":
                    status_info["item_class"] = "众筹成功"
                    status_info["is_success"] = True
                elif class_result_text == "项目终止":
                    status_info["item_class"] = "项目终止" # Often implies success or specific end state
                    status_info["is_success"] = True # Or a different flag if needed
                elif class_result_text == "众筹结束": # This usually means failed if not successful
                    status_info["item_class"] = "众筹失败"
                    status_info["is_fail"] = True
                    status_info["is_going"] = True # Was likely 'going' before ending as fail
                elif class_result_text == "众筹取消":
                    status_info["item_class"] = "众筹取消"
                    status_info["is_fail"] = True # Or a different flag
                    status_info["is_going"] = True # Was likely 'going' before cancel
        return status_info

    def parse_upper_items(self, soup_detail_page, project_status):
        """解析项目上部信息（时间、作者、基础资金信息）- 修复版"""
        data = []
        starttime = "none"
        endtime = "none"
        itemreal_class = project_status["item_class"]

        # 🔧 优化时间信息提取 - 从页面文本中提取
        page_text = soup_detail_page.get_text()

        # 提取开始时间
        start_time_patterns = [
            r'开始时间.*?(\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{2})',
            r'(\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{2}).*?开始',
            r'开始.*?(\d{4}-\d{1,2}-\d{1,2})',
            r'(\d{4}-\d{1,2}-\d{1,2}).*?开始'
        ]

        for pattern in start_time_patterns:
            start_match = re.search(pattern, page_text)
            if start_match:
                starttime = start_match.group(1)
                logger.info(f"找到开始时间: {starttime}")
                break

        # 提取结束时间
        end_time_patterns = [
            r'结束时间.*?(\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{2})',
            r'(\d{4}-\d{1,2}-\d{1,2}\s+\d{1,2}:\d{2}).*?结束',
            r'剩余时间.*?(\d{4}-\d{1,2}-\d{1,2})',
            r'(\d{4}-\d{1,2}-\d{1,2}).*?前达到'
        ]

        for pattern in end_time_patterns:
            end_match = re.search(pattern, page_text)
            if end_match:
                endtime = end_match.group(1)
                logger.info(f"找到结束时间: {endtime}")
                break

        # 根据项目状态设置默认时间
        if project_status["is_preheat"]:
            if starttime == "none":
                starttime = "预热中"
            if endtime == "none":
                endtime = "预热中"
        elif project_status["is_idea"]:
            starttime = "创意中"
            endtime = "创意中"

        data.extend([starttime, endtime, itemreal_class])

        # 🔧 优化作者信息提取 - 从页面文本中提取
        sponsor_href = "none"
        true_authorid_from_re = "none"
        author_image = "none"
        category = "none"
        author_name = "none"
        author_uid_attr = "0"
        parsed_author_page_details = ["0", "0", "0", "{}", "{}", "none"]

        # 从页面文本中提取作者名称 - 查找"发起了这个项目"前的文本
        author_match = re.search(r'([^\n]+)\s*发起了这个项目', page_text)
        if author_match:
            author_name = author_match.group(1).strip()
            logger.info(f"找到作者名称: {author_name}")

        # 从页面文本中提取项目分类 - "项目类别：桌游"
        category_match = re.search(r'项目类别[：:]\s*([^\n\r]+)', page_text)
        if category_match:
            category = category_match.group(1).strip()
            logger.info(f"找到项目分类: {category}")

        # 查找作者链接 - 查找包含uid的链接
        author_links = soup_detail_page.find_all('a', href=re.compile(r'uid=\d+'))
        if author_links:
            sponsor_href = author_links[0].get('href')
            if not sponsor_href.startswith('http'):
                sponsor_href = 'https://me.modian.com' + sponsor_href

            # 提取用户ID
            uid_match = re.search(r'uid=(\d+)', sponsor_href)
            if uid_match:
                true_authorid_from_re = uid_match.group(1)
                author_uid_attr = true_authorid_from_re
                logger.info(f"找到作者UID: {true_authorid_from_re}")

                # 获取作者页面信息
                try:
                    author_page_html = self.askURL2(sponsor_href)
                    if author_page_html:
                        parsed_author_page_details = self.parse_author_page_info(author_page_html, true_authorid_from_re)
                except Exception as e:
                    logger.warning(f"获取作者页面失败: {e}")

        # 查找作者头像
        author_imgs = soup_detail_page.find_all('img')
        for img in author_imgs:
            src = img.get('src')
            if src and ('avatar' in src or 'dst_avatar' in src):
                author_image = src
                if not author_image.startswith('http'):
                    author_image = 'https:' + author_image
                logger.info(f"找到作者头像: {author_image[:50]}...")
                break


        data.append(sponsor_href) # User homepage link
        data.append(author_image)
        data.append(category)
        data.append(author_name)
        data.append(author_uid_attr) # This is the one from data-username, often the same as true_authorid_from_re

        # 🔧 优化众筹信息提取 - 从页面文本中提取
        money = "0"
        percent = "0"
        goal_money = "0"
        sponsor_num = "0"

        # 从页面文本中提取已筹金额 - 处理编码问题
        money_patterns = [
            r'已筹[¥￥Â¥]([0-9,]+)',  # 正常编码
            r'å·²ç­¹[¥￥Â¥]([0-9,]+)',  # 编码后的中文
            r'已筹.*?[¥￥Â¥]\s*([0-9,]+)',  # 宽松匹配
            r'å·²ç­¹.*?[¥￥Â¥]\s*([0-9,]+)'   # 编码后宽松匹配
        ]

        for pattern in money_patterns:
            money_match = re.search(pattern, page_text)
            if money_match:
                money = money_match.group(1).replace(',', '')
                logger.info(f"找到已筹金额: ¥{money}")
                break

        # 从页面文本中提取目标金额
        goal_patterns = [
            r'目标金额\s*[¥￥Â¥]([0-9,]+)',  # 正常编码
            r'ç®æ éé¢\s*[¥￥Â¥]([0-9,]+)',  # 编码后的中文
            r'目标金额.*?[¥￥Â¥]\s*([0-9,]+)',  # 宽松匹配
            r'ç®æ éé¢.*?[¥￥Â¥]\s*([0-9,]+)',   # 编码后宽松匹配
            r'目标[¥￥Â¥]([0-9,]+)',  # 简化格式
            r'ç®æ[¥￥Â¥]([0-9,]+)',  # 编码后简化格式
            r'[¥￥Â¥]([0-9,]+).*?目标',  # 反向匹配
            r'([0-9,]+).*?目标金额'  # 数字在前
        ]

        for pattern in goal_patterns:
            goal_match = re.search(pattern, page_text)
            if goal_match:
                goal_money = goal_match.group(1).replace(',', '')
                logger.info(f"找到目标金额: ¥{goal_money}")
                break

        # 从页面文本中提取完成百分比
        percent_match = re.search(r'([0-9.]+)%', page_text)
        if percent_match:
            percent = percent_match.group(1)
            logger.info(f"找到完成百分比: {percent}%")

        # 如果没有找到目标金额，尝试从百分比反推
        if goal_money == "0" and money != "0" and percent != "0":
            try:
                calculated_goal = float(money) * 100 / float(percent)
                goal_money = str(int(calculated_goal))
                logger.info(f"从百分比反推目标金额: ¥{goal_money} (计算: {money} × 100 ÷ {percent})")
            except Exception as e:
                logger.warning(f"目标金额反推计算失败: {e}")

        # 从页面文本中提取支持者数量
        supporter_patterns = [
            r'(\d+)人\s*支持人数',  # 正常编码
            r'(\d+)äºº\s*æ¯æäººæ°',  # 编码后的中文
            r'支持人数\s*(\d+)',
            r'æ¯æäººæ°\s*(\d+)',
            r'(\d+)\s*人\s*支持',
            r'(\d+)\s*äºº\s*æ¯æ',
            r'支持者\s*(\d+)',
            r'æ¯æè\s*(\d+)',
            r'(\d+)\s*支持者',
            r'(\d+)\s*æ¯æè',
            r'(\d+)\s*人$',  # 简化格式
            r'(\d+)\s*äºº$'  # 编码后简化格式
        ]

        for pattern in supporter_patterns:
            supporter_match = re.search(pattern, page_text)
            if supporter_match:
                sponsor_num = supporter_match.group(1)
                logger.info(f"找到支持者数量: {sponsor_num}人")
                break

        # 验证数据合理性
        if money != "0" and goal_money != "0":
            try:
                calculated_percent = (float(money) / float(goal_money)) * 100
                if percent == "0":
                    percent = f"{calculated_percent:.1f}"
            except:
                pass

        logger.info(f"解析众筹信息: 已筹¥{money}, 目标¥{goal_money}, 完成率{percent}%, 支持者{sponsor_num}人")

        data.extend([money, percent, goal_money, sponsor_num])
        data.append(true_authorid_from_re) # The ID extracted from sponsor_href
        data.extend(parsed_author_page_details) # fans_num, notice_number, love_number, detail_result_str, other_result_str, userhome_url_confirmation
        return data

    def parse_main_left_content(self, soup_detail_page):
        """解析项目左侧内容（图片和视频）- 修复版"""
        data = []
        img_list = []
        video_list = []

        # 🔧 优化媒体内容提取 - 查找所有相关区域
        content_areas = [
            soup_detail_page.find('div', {'class': 'main-left'}),
            soup_detail_page.find('div', {'class': 'project-content'}),
            soup_detail_page.find('div', {'class': 'content-wrap'}),
            soup_detail_page.find('div', {'id': 'projectDetail'}),
            soup_detail_page.find('section', {'class': 'project-detail'})
        ]

        for area in content_areas:
            if area:
                # 查找图片
                for img_tag in area.find_all('img'):
                    src = img_tag.get('src')
                    if src and src.strip():
                        # 过滤掉头像、图标等无关图片
                        if not any(keyword in src for keyword in [
                            'default_profile', 'icon-', 'avatar', 'logo',
                            'headPic', 'default_1x1', 'video-play'
                        ]):
                            # 确保URL完整
                            if src.startswith('//'):
                                src = 'https:' + src
                            elif src.startswith('/'):
                                src = 'https://zhongchou.modian.com' + src
                            img_list.append(src)

                # 查找视频
                for video_tag in area.find_all('video'):
                    if video_tag.get('src'):
                        video_src = video_tag.get('src')
                        if video_src.startswith('//'):
                            video_src = 'https:' + video_src
                        elif video_src.startswith('/'):
                            video_src = 'https://zhongchou.modian.com' + video_src
                        video_list.append(video_src)
                    else:
                        # 查找source标签
                        for source_tag in video_tag.find_all('source'):
                            if source_tag.get('src'):
                                video_src = source_tag.get('src')
                                if video_src.startswith('//'):
                                    video_src = 'https:' + video_src
                                elif video_src.startswith('/'):
                                    video_src = 'https://zhongchou.modian.com' + video_src
                                video_list.append(video_src)
                                break

                # 查找iframe中的视频（如B站、优酷等）
                for iframe_tag in area.find_all('iframe'):
                    src = iframe_tag.get('src')
                    if src and any(domain in src for domain in [
                        'bilibili', 'youku', 'qq.com', 'iqiyi', 'tudou', 'youtube'
                    ]):
                        if src.startswith('//'):
                            src = 'https:' + src
                        video_list.append(src)

        # 🔧 从页面文本中查找视频链接
        page_text = soup_detail_page.get_text()
        video_url_patterns = [
            r'https?://[^\s]*\.mp4',
            r'https?://[^\s]*\.avi',
            r'https?://[^\s]*\.mov',
            r'https?://[^\s]*\.wmv',
            r'https?://mediac\d+\.modian\.com/[^\s]*\.mp4'
        ]

        for pattern in video_url_patterns:
            video_matches = re.findall(pattern, page_text)
            for match in video_matches:
                video_list.append(match)

        # 去重并过滤
        img_list = list(set([img for img in img_list if img and len(img) > 10]))
        video_list = list(set([video for video in video_list if video and len(video) > 10]))

        logger.info(f"找到项目媒体: {len(img_list)}张图片, {len(video_list)}个视频")

        data.extend([len(img_list), str(img_list), len(video_list), str(video_list)])
        return data

    def parse_main_right_rewards(self, soup_detail_page):
        """解析项目右侧回报信息"""
        rewards_data_list_str = [] # List of strings, each representing a reward tier
        main_right_div = soup_detail_page.find('div', {'class': 'main-right'})
        if main_right_div:
            payback_lists_div = main_right_div.find('div', {'class': 'payback-lists margin36'})
            if payback_lists_div:
                for reward_item_div in payback_lists_div.find_all('div', class_=lambda x: x and 'back-list' in x):
                    single_reward_details = []

                    back_money = "0"
                    head_div = reward_item_div.find('div', {'class': 'head'})
                    if head_div and head_div.find('span'):
                        back_money_text = head_div.find('span').text.strip().replace('￥','')
                        if back_money_text.isdigit(): back_money = back_money_text

                    backsponsor = "0"
                    if head_div and head_div.find('em'):
                        em_text = head_div.find('em').text.strip() # e.g., "23 人支持" or "已满"
                        parts = em_text.split()
                        if parts and parts[0].isdigit(): backsponsor = parts[0]
                        elif "已满" in em_text: backsponsor = "已满"


                    sign_logo = "0" # Default, might be "限量" or number
                    zc_subhead_div = reward_item_div.find('div', {'class': 'zc-subhead'})
                    if zc_subhead_div and zc_subhead_div.find('span'):
                        sign_logo_text = zc_subhead_div.find('span').text.strip()
                        if "限量" in sign_logo_text:
                            num_part = sign_logo_text.replace("限量","").replace("份","").strip()
                            sign_logo = f"限量 {num_part}" if num_part.isdigit() else "限量"
                        elif sign_logo_text.isdigit(): # Unlikely based on old regex, but good to check
                            sign_logo = sign_logo_text


                    backtitle = "none"
                    backtext = "none"
                    backtime = "none"

                    back_content_div = reward_item_div.find('div', {'class': 'back-content'})
                    if back_content_div:
                        title_div = back_content_div.find('div', {'class': 'back-sub-title'})
                        if title_div: backtitle = title_div.text.strip()

                        detail_div = back_content_div.find('div', {'class': 'back-detail'})
                        if detail_div: backtext = detail_div.text.strip()

                        time_div = back_content_div.find('div', {'class': 'back-time'})
                        if time_div: backtime = time_div.text.strip()

                    single_reward_details.extend([backtitle, sign_logo, back_money, backsponsor, backtime, backtext])
                    rewards_data_list_str.append(str(single_reward_details))

        return [str(rewards_data_list_str), len(rewards_data_list_str)]

    def parse_main_middle_nav_info(self, soup_detail_page, project_status):
        """解析项目中间导航信息（更新数、评论数、支持者数、收藏数）- 深度修复版"""
        data = []
        update_number = "0"
        comment_number = "0"
        userlist_num = "0"
        collect_number = "0"

        # 🔧 优化策略：从页面文本中直接提取数字信息
        page_text = soup_detail_page.get_text()

        # 提取更新数 - 查找"项目更新 X"模式
        update_patterns = [
            r'项目更新\s*(\d+)',
            r'更新\s*(\d+)',
            r'(\d+)\s*次更新',
            r'更新.*?(\d+)'
        ]

        for pattern in update_patterns:
            update_match = re.search(pattern, page_text)
            if update_match:
                update_number = update_match.group(1)
                logger.info(f"找到更新数: {update_number}")
                break

        # 提取评论数 - 查找"评论 X"模式
        comment_patterns = [
            r'评论\s*(\d+)',
            r'(\d+)\s*条评论',
            r'评论.*?(\d+)',
            r'(\d+)\s*评论'
        ]

        for pattern in comment_patterns:
            comment_match = re.search(pattern, page_text)
            if comment_match:
                comment_number = comment_match.group(1)
                logger.info(f"找到评论数: {comment_number}")
                break

        # 提取支持者数 - 查找"支持者 X"或"X人"模式
        supporter_patterns = [
            r'支持者\s*(\d+)',
            r'(\d+)\s*人\s*支持',
            r'(\d+)\s*支持者',
            r'支持人数.*?(\d+)',
            r'(\d+)\s*人$'  # 行末的数字+人
        ]

        for pattern in supporter_patterns:
            supporter_match = re.search(pattern, page_text)
            if supporter_match:
                userlist_num = supporter_match.group(1)
                logger.info(f"找到支持者数: {userlist_num}")
                break

        # 提取收藏数 - 查找收藏相关数字
        collect_patterns = [
            r'收藏\s*(\d+)',
            r'(\d+)\s*收藏',
            r'关注\s*(\d+)',
            r'(\d+)\s*关注'
        ]

        for pattern in collect_patterns:
            collect_match = re.search(pattern, page_text)
            if collect_match:
                collect_number = collect_match.group(1)
                logger.info(f"找到收藏数: {collect_number}")
                break

        # 🔧 回退到传统DOM解析（如果文本解析失败）
        if all(x == "0" for x in [update_number, comment_number, userlist_num, collect_number]):
            logger.info("文本解析失败，回退到DOM解析")
            nav_wrap_inner = soup_detail_page.find('div', {'class': 'nav-wrap-inner'})
            if nav_wrap_inner:
                nav_left = nav_wrap_inner.find('ul', {'class': 'nav-left'})
                if nav_left:
                    # 更新数
                    update_li = nav_left.find('li', {'class': 'pro-gengxin'})
                    if update_li:
                        li_text = update_li.get_text()
                        numbers = re.findall(r'\d+', li_text)
                        if numbers:
                            update_number = numbers[-1]

                    # 评论数
                    comment_li = nav_left.find('li', {'class': 'nav-comment'})
                    if comment_li:
                        li_text = comment_li.get_text()
                        numbers = re.findall(r'\d+', li_text)
                        if numbers:
                            comment_number = numbers[-1]

                    # 支持者数
                    userlist_li = nav_left.find('li', class_='dialog_user_list')
                    if userlist_li:
                        li_text = userlist_li.get_text()
                        numbers = re.findall(r'\d+', li_text)
                        if numbers:
                            userlist_num = numbers[-1]

                # 收藏数
                nav_right = nav_wrap_inner.find('ul', {'class': 'nav-right'})
                if nav_right:
                    atten_li = nav_right.find('li', {'class': 'atten'})
                    if atten_li:
                        li_text = atten_li.get_text()
                        numbers = re.findall(r'\d+', li_text)
                        if numbers:
                            collect_number = numbers[-1]

        # 🔧 数据清理和验证
        def clean_number(num_str, field_name=""):
            """增强的数字清理函数"""
            if not num_str:
                return "0"

            num_str = str(num_str).strip()
            cleaned = re.sub(r'[^\d]', '', num_str)

            if cleaned and cleaned.isdigit():
                return cleaned
            else:
                return "0"

        update_number = clean_number(update_number, "更新数")
        comment_number = clean_number(comment_number, "评论数")
        userlist_num = clean_number(userlist_num, "支持者数")
        collect_number = clean_number(collect_number, "收藏数")

        logger.info(f"导航信息解析结果: 更新数={update_number}, 评论数={comment_number}, 支持者数={userlist_num}, 收藏数={collect_number}")

        data.extend([update_number, comment_number, userlist_num, collect_number])
        return data

    def parse_project_detail_page(self, html_content):
        """解析项目详情页面"""
        soup = BeautifulSoup(html_content, "html.parser")
        project_data = []

        project_status = self.get_project_status_info(soup)

        # Upper items: time, author, basic funding info
        upper_items_data = self.parse_upper_items(soup, project_status)
        project_data.extend(upper_items_data)

        # Right items: Rewards
        main_right_data = self.parse_main_right_rewards(soup)
        project_data.extend(main_right_data)

        # Middle items: Nav counts (updates, comments, supporters/likes, collections)
        main_middle_data = self.parse_main_middle_nav_info(soup, project_status)
        project_data.extend(main_middle_data)

        # Left items: Project content images/videos
        main_left_data = self.parse_main_left_content(soup)
        project_data.extend(main_left_data)

        return project_data

    def parse_main_listing_page(self, html_content, current_excel_index_ref):
        """解析主列表页面"""
        soup = BeautifulSoup(html_content, "html.parser")
        datalist_batch = []

        pro_field_div = soup.find('div', {'class': 'pro_field'})
        if not pro_field_div:
            logger.warning("No 'pro_field' div found on listing page.")
            return datalist_batch, current_excel_index_ref

        # Find all project items - they are in <li> elements within the pro_field div
        project_items = pro_field_div.find_all('li')
        if not project_items:
            logger.warning("No project items found in pro_field div.")
            return datalist_batch, current_excel_index_ref

        for item_li in project_items:
            current_excel_index_ref += 1 # Increment shared index
            single_project_base_data = [current_excel_index_ref] # Start with new auto-incremented index

            # Find the project link - look for any <a> tag with href containing "/item/"
            link_tag = None
            item_link = "none"
            item_id = ""

            # Try to find the main project link
            for a_tag in item_li.find_all('a'):
                href = a_tag.get('href', '')
                if '/item/' in href:
                    link_tag = a_tag
                    item_link = href
                    break

            if link_tag and item_link != "none":
                if not item_link.startswith("http"):
                    item_link = "https://zhongchou.modian.com" + item_link

                id_match = self.config.LINK_ID_PATTERN.search(item_link)
                if id_match:
                    item_id = id_match.group(1)

            single_project_base_data.append(item_link)
            single_project_base_data.append(item_id)

            # Find the project title
            title = "none"
            title_h3 = item_li.find('h3', class_='pro_title')
            if title_h3:
                title = title_h3.text.strip()
            elif link_tag and link_tag.text:
                title = link_tag.text.strip()

            # Skip specific items as per original logic
            if "可汗游戏大会" in title:
                current_excel_index_ref -= 1 # Decrement back as we are skipping
                continue

            single_project_base_data.append(title)

            # Find the project image
            img_src = "none"
            img_tag = item_li.find('img')
            if img_tag and img_tag.get('src'):
                img_src = img_tag.get('src')
            single_project_base_data.append(img_src)

            logger.info(f"Processing: {current_excel_index_ref} - {title} ({item_link})")

            # Fetch and parse detail page
            if item_link != "none" and item_id: # Only proceed if we have a valid link/ID
                detail_page_html = self.askURL(item_link)
                if detail_page_html:
                    detail_page_data = self.parse_project_detail_page(detail_page_html)
                    single_project_base_data.extend(detail_page_data)
                    self.stats.projects_processed += 1
                else:
                    logger.warning(f"Failed to fetch detail page for {item_link}")
                    # Add placeholders for detail_data if fetch fails, to maintain column consistency
                    # Number of placeholders should match expected fields from parse_project_detail_page
                    # upper (11) + right (2) + middle (4) + left (4) = 21 placeholders
                    single_project_base_data.extend(["error_fetching_details"] * 21)
            else:
                logger.warning(f"Skipping detail fetch for item without link/ID: {title}")
                single_project_base_data.extend(["no_link_for_details"] * 21)

            datalist_batch.append(single_project_base_data)

        return datalist_batch, current_excel_index_ref

    def save_data_to_excel(self, workbook, sheet, data_rows_list, start_row_idx):
        """保存数据到Excel文件"""
        logger.info(f"Saving {len(data_rows_list)} rows to Excel, starting at sheet row {start_row_idx + 1}...")

        col_headers = (
            "序号", "项目link", "项目6位id", "项目名称", "项目图",  # Base info from listing (5)
            "开始时间", "结束时间", "项目结果",  # from upper (3)
            "用户主页(链接)", "用户头像(图片链接)", "分类", "用户名", "用户UID(data-username)", # from upper (5)
            "已筹金额", "百分比", "目标金额", "支持者(数量)", # from upper (4)
            "真实用户ID(链接提取)", "作者页-粉丝数", "作者页-关注数", "作者页-获赞数", "作者页-详情", "作者页-其他信息", "作者页-主页确认", # from upper's author parse (7)
            "回报列表信息(字符串)", "回报列表项目数",  # from right (2)
            "项目更新数", "评论数", "项目支持者/点赞数", "收藏数",  # from middle (4)
            "项目详情-图片数量", "项目详情-图片(列表字符串)", "项目详情-视频数量", "项目详情-视频(列表字符串)"  # from left (4)
        ) # Total 34 columns

        # Write headers only once
        if start_row_idx == 0:
            for i, header_name in enumerate(col_headers):
                sheet.write(0, i, header_name)

        # Process each row of data
        for row_idx, row_data in enumerate(data_rows_list):
            if not row_data:
                continue

            # Calculate the actual Excel row number
            excel_row_num = start_row_idx + row_idx + 1  # +1 because headers are at row 0

            # Ensure row_data is a list and has enough elements
            if not isinstance(row_data, list):
                row_data = list(row_data) if hasattr(row_data, '__iter__') else [row_data]

            # Pad with empty strings if not enough columns
            padded_row_data = row_data + [""] * (len(col_headers) - len(row_data))

            # Write each cell in the row
            for col_idx, cell_value in enumerate(padded_row_data):
                cell_str = str(cell_value) if cell_value is not None else ""

                # Handle Excel cell character limit
                if len(cell_str) > 32767:
                    cell_str = cell_str[:32767] + "...TRUNCATED"

                try:
                    sheet.write(excel_row_num, col_idx, cell_str)
                except Exception as e:
                    logger.error(f"Error writing to Excel cell ({excel_row_num},{col_idx}): {e}. Value: {cell_str[:50]}")
                    try:
                        sheet.write(excel_row_num, col_idx, "ERROR_WRITING_CELL")
                    except:
                        pass  # Skip if even error message can't be written

        # Save the workbook
        try:
            excel_path = self.config.OUTPUT_DIR / self.config.EXCEL_FILENAME
            workbook.save(excel_path)
            logger.info(f"Data saved to {excel_path}")
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")

    def save_data_to_multiple_formats(self, projects_data: List[List[Any]], suffix: str = "",
                                     stats: Optional[Dict[str, Any]] = None):
        """保存数据到多种格式（增强版）"""
        if not projects_data:
            logger.warning("没有数据需要保存")
            return {}

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"modian_projects_{timestamp}"
        if suffix:
            base_filename += f"_{suffix}"

        saved_files = {}

        # 数据质量分析
        quality_report = self._analyze_data_quality(projects_data)

        try:
            # 1. 保存为增强JSON格式（优化AI分析）
            json_file = self._save_enhanced_json(projects_data, base_filename, quality_report, stats)
            saved_files['json'] = json_file

            # 2. 保存为CSV格式
            csv_file = self._save_enhanced_csv(projects_data, base_filename)
            saved_files['csv'] = csv_file

            # 3. 保存为Excel格式（兼容原有格式）
            excel_file = self._save_enhanced_excel(projects_data, base_filename)
            saved_files['excel'] = excel_file

            # 4. 生成数据质量报告
            report_file = self._save_quality_report(quality_report, base_filename, stats)
            saved_files['quality_report'] = report_file

            logger.info(f"✅ 数据已保存为多种格式:")
            for format_type, file_path in saved_files.items():
                logger.info(f"  {format_type.upper()}: {file_path}")

            return saved_files

        except Exception as e:
            logger.error(f"保存多格式文件失败: {e}")
            return saved_files

    def _analyze_data_quality(self, projects_data: List[List[Any]]) -> Dict[str, Any]:
        """分析数据质量"""
        if not projects_data:
            return {"total_projects": 0, "empty_fields": {}, "data_completeness": 0}

        # 定义字段名称（与Excel列对应）
        field_names = [
            "序号", "项目link", "项目6位id", "项目名称", "项目图",
            "开始时间", "结束时间", "项目结果",
            "用户主页(链接)", "用户头像(图片链接)", "分类", "用户名", "用户UID(data-username)",
            "已筹金额", "百分比", "目标金额", "支持者(数量)",
            "真实用户ID(链接提取)", "作者页-粉丝数", "作者页-关注数", "作者页-获赞数",
            "作者页-详情", "作者页-其他信息", "作者页-主页确认",
            "回报列表信息(字符串)", "回报列表项目数",
            "项目更新数", "评论数", "项目支持者/点赞数", "收藏数",
            "项目详情-图片数量", "项目详情-图片(列表字符串)",
            "项目详情-视频数量", "项目详情-视频(列表字符串)"
        ]

        total_projects = len(projects_data)
        empty_fields = {}
        error_fields = {}

        # 分析每个字段的数据质量
        for field_idx, field_name in enumerate(field_names):
            empty_count = 0
            error_count = 0

            for project in projects_data:
                if field_idx >= len(project):
                    empty_count += 1
                else:
                    value = project[field_idx]
                    if value is None or str(value).strip() in ["", "none", "0", "error_fetching_details", "no_link_for_details"]:
                        empty_count += 1
                    elif "error" in str(value).lower():
                        error_count += 1

            if empty_count > 0:
                empty_fields[field_name] = {
                    "empty_count": empty_count,
                    "empty_percentage": round((empty_count / total_projects) * 100, 2)
                }

            if error_count > 0:
                error_fields[field_name] = {
                    "error_count": error_count,
                    "error_percentage": round((error_count / total_projects) * 100, 2)
                }

        # 计算整体数据完整性
        total_fields = len(field_names) * total_projects
        total_empty = sum(field["empty_count"] for field in empty_fields.values())
        data_completeness = round(((total_fields - total_empty) / total_fields) * 100, 2) if total_fields > 0 else 0

        return {
            "total_projects": total_projects,
            "total_fields": len(field_names),
            "empty_fields": empty_fields,
            "error_fields": error_fields,
            "data_completeness": data_completeness,
            "analysis_time": datetime.datetime.now().isoformat()
        }

    def _save_enhanced_json(self, projects_data: List[List[Any]], base_filename: str,
                           quality_report: Dict[str, Any], stats: Optional[Dict[str, Any]] = None) -> str:
        """保存增强的JSON格式（优化AI分析）"""
        json_file = self.config.OUTPUT_DIR / f"{base_filename}.json"

        # 字段名称映射
        field_names = [
            "序号", "项目link", "项目6位id", "项目名称", "项目图",
            "开始时间", "结束时间", "项目结果",
            "用户主页(链接)", "用户头像(图片链接)", "分类", "用户名", "用户UID(data-username)",
            "已筹金额", "百分比", "目标金额", "支持者(数量)",
            "真实用户ID(链接提取)", "作者页-粉丝数", "作者页-关注数", "作者页-获赞数",
            "作者页-详情", "作者页-其他信息", "作者页-主页确认",
            "回报列表信息(字符串)", "回报列表项目数",
            "项目更新数", "评论数", "项目支持者/点赞数", "收藏数",
            "项目详情-图片数量", "项目详情-图片(列表字符串)",
            "项目详情-视频数量", "项目详情-视频(列表字符串)"
        ]

        # 转换为结构化JSON
        projects_json = []
        for project_data in projects_data:
            project_dict = {}

            # 基本信息
            project_dict["basic_info"] = {
                "id": project_data[0] if len(project_data) > 0 else "",
                "link": project_data[1] if len(project_data) > 1 else "",
                "project_id": project_data[2] if len(project_data) > 2 else "",
                "title": project_data[3] if len(project_data) > 3 else "",
                "image": project_data[4] if len(project_data) > 4 else ""
            }

            # 时间信息
            project_dict["time_info"] = {
                "start_time": project_data[5] if len(project_data) > 5 else "",
                "end_time": project_data[6] if len(project_data) > 6 else "",
                "status": project_data[7] if len(project_data) > 7 else ""
            }

            # 作者信息
            project_dict["author_info"] = {
                "homepage": project_data[8] if len(project_data) > 8 else "",
                "avatar": project_data[9] if len(project_data) > 9 else "",
                "category": project_data[10] if len(project_data) > 10 else "",
                "name": project_data[11] if len(project_data) > 11 else "",
                "uid": project_data[12] if len(project_data) > 12 else "",
                "real_id": project_data[17] if len(project_data) > 17 else "",
                "fans_count": self._safe_int(project_data[18] if len(project_data) > 18 else 0),
                "following_count": self._safe_int(project_data[19] if len(project_data) > 19 else 0),
                "likes_count": self._safe_int(project_data[20] if len(project_data) > 20 else 0),
                "details": project_data[21] if len(project_data) > 21 else "",
                "other_info": project_data[22] if len(project_data) > 22 else "",
                "homepage_confirm": project_data[23] if len(project_data) > 23 else ""
            }

            # 众筹信息
            project_dict["funding_info"] = {
                "raised_amount": self._safe_float(project_data[13] if len(project_data) > 13 else 0),
                "completion_rate": self._safe_float(project_data[14] if len(project_data) > 14 else 0),
                "target_amount": self._safe_float(project_data[15] if len(project_data) > 15 else 0),
                "backer_count": self._safe_int(project_data[16] if len(project_data) > 16 else 0)
            }

            # 回报信息
            project_dict["reward_info"] = {
                "rewards_list": project_data[24] if len(project_data) > 24 else "",
                "rewards_count": self._safe_int(project_data[25] if len(project_data) > 25 else 0)
            }

            # 互动信息
            project_dict["engagement_info"] = {
                "updates_count": self._safe_int(project_data[26] if len(project_data) > 26 else 0),
                "comments_count": self._safe_int(project_data[27] if len(project_data) > 27 else 0),
                "supporters_likes_count": self._safe_int(project_data[28] if len(project_data) > 28 else 0),
                "collections_count": self._safe_int(project_data[29] if len(project_data) > 29 else 0)
            }

            # 内容信息
            project_dict["content_info"] = {
                "images_count": self._safe_int(project_data[30] if len(project_data) > 30 else 0),
                "images_list": project_data[31] if len(project_data) > 31 else "",
                "videos_count": self._safe_int(project_data[32] if len(project_data) > 32 else 0),
                "videos_list": project_data[33] if len(project_data) > 33 else ""
            }

            projects_json.append(project_dict)

        # 构建完整的JSON结构
        json_data = {
            "metadata": {
                "export_time": datetime.datetime.now().isoformat(),
                "total_projects": len(projects_json),
                "format_version": "2.0",
                "data_source": "modian_crowdfunding",
                "fields_count": len(field_names),
                "quality_report": quality_report
            },
            "statistics": stats if stats else {},
            "projects": projects_json
        }

        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        logger.info(f"📄 增强JSON文件已保存: {json_file}")
        return str(json_file)

    def _safe_int(self, value) -> int:
        """安全转换为整数"""
        try:
            if isinstance(value, str):
                # 移除常见的非数字字符
                cleaned = value.replace(',', '').replace('人支持', '').replace('人订阅', '').strip()
                if cleaned.isdigit():
                    return int(cleaned)
            elif isinstance(value, (int, float)):
                return int(value)
            return 0
        except (ValueError, TypeError):
            return 0

    def _safe_float(self, value) -> float:
        """安全转换为浮点数"""
        try:
            if isinstance(value, str):
                # 移除常见的非数字字符
                cleaned = value.replace(',', '').replace('￥', '').replace('%', '').strip()
                if cleaned.replace('.', '').isdigit():
                    return float(cleaned)
            elif isinstance(value, (int, float)):
                return float(value)
            return 0.0
        except (ValueError, TypeError):
            return 0.0

    def _save_enhanced_csv(self, projects_data: List[List[Any]], base_filename: str) -> str:
        """保存增强的CSV格式"""
        csv_file = self.config.OUTPUT_DIR / f"{base_filename}.csv"

        # 字段名称
        field_names = [
            "序号", "项目link", "项目6位id", "项目名称", "项目图",
            "开始时间", "结束时间", "项目结果",
            "用户主页(链接)", "用户头像(图片链接)", "分类", "用户名", "用户UID(data-username)",
            "已筹金额", "百分比", "目标金额", "支持者(数量)",
            "真实用户ID(链接提取)", "作者页-粉丝数", "作者页-关注数", "作者页-获赞数",
            "作者页-详情", "作者页-其他信息", "作者页-主页确认",
            "回报列表信息(字符串)", "回报列表项目数",
            "项目更新数", "评论数", "项目支持者/点赞数", "收藏数",
            "项目详情-图片数量", "项目详情-图片(列表字符串)",
            "项目详情-视频数量", "项目详情-视频(列表字符串)"
        ]

        try:
            with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)

                # 写入表头
                writer.writerow(field_names)

                # 写入数据
                for project_data in projects_data:
                    # 确保数据长度匹配表头，填充空值
                    padded_data = list(project_data) + [""] * (len(field_names) - len(project_data))

                    # 清理数据
                    cleaned_row = []
                    for i, cell in enumerate(padded_data[:len(field_names)]):
                        cell_str = str(cell) if cell is not None else ""
                        # 清理CSV中的特殊字符
                        cell_str = cell_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                        cleaned_row.append(cell_str)

                    writer.writerow(cleaned_row)

            logger.info(f"📊 增强CSV文件已保存: {csv_file}")
            return str(csv_file)

        except Exception as e:
            logger.error(f"保存CSV文件失败: {e}")
            raise

    def _save_enhanced_excel(self, projects_data: List[List[Any]], base_filename: str) -> str:
        """保存增强的Excel格式（兼容原有格式）"""
        excel_file = self.config.OUTPUT_DIR / f"{base_filename}.xlsx"

        try:
            # 使用openpyxl创建工作簿
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "摩点众筹项目数据"

            # 字段名称
            field_names = [
                "序号", "项目link", "项目6位id", "项目名称", "项目图",
                "开始时间", "结束时间", "项目结果",
                "用户主页(链接)", "用户头像(图片链接)", "分类", "用户名", "用户UID(data-username)",
                "已筹金额", "百分比", "目标金额", "支持者(数量)",
                "真实用户ID(链接提取)", "作者页-粉丝数", "作者页-关注数", "作者页-获赞数",
                "作者页-详情", "作者页-其他信息", "作者页-主页确认",
                "回报列表信息(字符串)", "回报列表项目数",
                "项目更新数", "评论数", "项目支持者/点赞数", "收藏数",
                "项目详情-图片数量", "项目详情-图片(列表字符串)",
                "项目详情-视频数量", "项目详情-视频(列表字符串)"
            ]

            # 写入表头
            for col_idx, header in enumerate(field_names, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, project_data in enumerate(projects_data, 2):
                for col_idx, cell_value in enumerate(project_data, 1):
                    if col_idx > len(field_names):
                        break

                    # 处理单元格值
                    if cell_value is None:
                        cell_value = ""
                    else:
                        cell_str = str(cell_value)
                        # Excel单元格字符限制
                        if len(cell_str) > 32767:
                            cell_str = cell_str[:32764] + "..."
                        cell_value = cell_str

                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # 保存文件
            wb.save(excel_file)
            logger.info(f"📈 增强Excel文件已保存: {excel_file}")
            return str(excel_file)

        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            # 回退到xlwt格式
            try:
                excel_file_xls = self.config.OUTPUT_DIR / f"{base_filename}.xls"
                workbook = xlwt.Workbook(encoding="utf-8", style_compression=0)
                sheet = workbook.add_sheet('projects', cell_overwrite_ok=True)

                # 使用原有的保存方法
                self.save_data_to_excel(workbook, sheet, projects_data, 0)
                logger.info(f"📈 Excel文件已保存(XLS格式): {excel_file_xls}")
                return str(excel_file_xls)
            except Exception as e2:
                logger.error(f"保存XLS文件也失败: {e2}")
                raise

    def _save_quality_report(self, quality_report: Dict[str, Any], base_filename: str,
                           stats: Optional[Dict[str, Any]] = None) -> str:
        """保存数据质量报告"""
        report_file = self.config.OUTPUT_DIR / f"{base_filename}_quality_report.txt"

        try:
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("摩点众筹爬虫 - 数据质量分析报告\n")
                f.write("=" * 80 + "\n\n")

                # 基本信息
                f.write(f"📊 报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"📈 数据项目总数: {quality_report.get('total_projects', 0)}\n")
                f.write(f"📋 数据字段总数: {quality_report.get('total_fields', 0)}\n")
                f.write(f"✅ 数据完整性: {quality_report.get('data_completeness', 0)}%\n\n")

                # 爬虫统计信息
                if stats:
                    f.write("🚀 爬虫运行统计:\n")
                    f.write("-" * 40 + "\n")
                    f.write(f"  运行时间: {stats.get('elapsed_time', 0):.2f} 秒\n")
                    f.write(f"  处理页面: {stats.get('pages_processed', 0)}\n")
                    f.write(f"  总请求数: {stats.get('total_requests', 0)}\n")
                    f.write(f"  成功请求: {stats.get('successful_requests', 0)}\n")
                    f.write(f"  失败请求: {stats.get('failed_requests', 0)}\n")
                    f.write(f"  成功率: {stats.get('success_rate', 0):.1f}%\n")
                    f.write(f"  发现项目: {stats.get('projects_found', 0)}\n")
                    f.write(f"  处理项目: {stats.get('projects_processed', 0)}\n\n")

                # 数据缺失分析
                empty_fields = quality_report.get('empty_fields', {})
                if empty_fields:
                    f.write("⚠️  数据缺失字段分析:\n")
                    f.write("-" * 40 + "\n")

                    # 按缺失率排序
                    sorted_empty = sorted(empty_fields.items(),
                                        key=lambda x: x[1]['empty_percentage'], reverse=True)

                    for field_name, field_info in sorted_empty:
                        f.write(f"  {field_name}:\n")
                        f.write(f"    缺失数量: {field_info['empty_count']}\n")
                        f.write(f"    缺失率: {field_info['empty_percentage']}%\n\n")
                else:
                    f.write("✅ 所有字段数据完整，无缺失数据\n\n")

                # 错误字段分析
                error_fields = quality_report.get('error_fields', {})
                if error_fields:
                    f.write("❌ 数据错误字段分析:\n")
                    f.write("-" * 40 + "\n")

                    for field_name, field_info in error_fields.items():
                        f.write(f"  {field_name}:\n")
                        f.write(f"    错误数量: {field_info['error_count']}\n")
                        f.write(f"    错误率: {field_info['error_percentage']}%\n\n")
                else:
                    f.write("✅ 无数据错误字段\n\n")

                # 数据质量建议
                f.write("💡 数据质量改进建议:\n")
                f.write("-" * 40 + "\n")

                completeness = quality_report.get('data_completeness', 0)
                if completeness >= 90:
                    f.write("  ✅ 数据质量优秀，完整性超过90%\n")
                elif completeness >= 80:
                    f.write("  ⚠️  数据质量良好，建议优化缺失率较高的字段\n")
                elif completeness >= 70:
                    f.write("  ⚠️  数据质量一般，需要重点优化数据采集逻辑\n")
                else:
                    f.write("  ❌ 数据质量较差，建议全面检查爬虫逻辑\n")

                # 针对高缺失率字段的建议
                high_missing_fields = [name for name, info in empty_fields.items()
                                     if info['empty_percentage'] > 50]
                if high_missing_fields:
                    f.write(f"\n  🔧 高缺失率字段({len(high_missing_fields)}个)需要优化:\n")
                    for field in high_missing_fields[:5]:  # 只显示前5个
                        f.write(f"    - {field}\n")

                f.write("\n" + "=" * 80 + "\n")
                f.write("报告结束\n")

            logger.info(f"📋 数据质量报告已保存: {report_file}")
            return str(report_file)

        except Exception as e:
            logger.error(f"保存质量报告失败: {e}")
            raise

    def print_stats(self):
        """打印统计信息"""
        stats = self.stats.get_summary()

        print("\n" + "="*60)
        print("📊 爬虫运行统计")
        print("="*60)
        print(f"运行时间: {stats['elapsed_time']:.2f} 秒")
        print(f"处理页面: {stats['pages_processed']}")
        print(f"总请求数: {stats['total_requests']}")
        print(f"成功请求: {stats['successful_requests']}")
        print(f"失败请求: {stats['failed_requests']}")
        print(f"成功率: {stats['success_rate']:.1f}%")
        print(f"发现项目: {stats['projects_found']}")
        print(f"处理项目: {stats['projects_processed']}")
        print(f"平均每页耗时: {stats['avg_time_per_page']:.2f} 秒")

        if stats['errors']:
            print("\n错误统计:")
            for error_type, messages in stats['errors'].items():
                print(f"  {error_type}: {len(messages)} 次")

        print("="*60)

    def run_scraper(self):
        """运行完整的爬虫流程"""
        logger.info("开始运行摩点众筹爬虫")

        # 创建Excel工作簿
        main_workbook = xlwt.Workbook(encoding="utf-8", style_compression=0)
        main_sheet = main_workbook.add_sheet('all_projects', cell_overwrite_ok=True)

        # 全局Excel行索引计数器
        excel_index_counter = 0 # Start at 0 because headers are row 0, data starts at row 1

        # 写入表头
        self.save_data_to_excel(main_workbook, main_sheet, [], 0) # Pass empty data list to just write headers

        # 爬取数据
        total_projects_processed = 0

        try:
            for page_num in range(1, self.config.MAX_PAGES + 1):
                logger.info(f"\n--- 正在处理第 {page_num} 页 ---")
                current_page_url = self.config.get_page_url(page_num)
                page_html = self.askURL(current_page_url)

                if not page_html:
                    logger.warning(f"Failed to get HTML for page {page_num}. Skipping.")
                    continue

                # 解析页面项目
                projects_on_page, excel_index_counter = self.parse_main_listing_page(page_html, excel_index_counter)

                if projects_on_page:
                    # 保存到Excel
                    self.save_data_to_excel(main_workbook, main_sheet, projects_on_page, 0)
                    total_projects_processed += len(projects_on_page)
                    logger.info(f"Processed {len(projects_on_page)} projects from page {page_num}. Total processed: {total_projects_processed}")

                    # 更新统计
                    self.stats.projects_found += len(projects_on_page)
                else:
                    logger.warning(f"No projects found or processed on page {page_num}.")

                self.stats.pages_processed += 1

                # 定期保存
                if page_num % self.config.SAVE_INTERVAL == 0:
                    logger.info(f"Intermediate save at page {page_num}...")
                    excel_path = self.config.OUTPUT_DIR / self.config.EXCEL_FILENAME
                    main_workbook.save(excel_path)

                # 添加延迟避免过于频繁的请求
                time.sleep(random.uniform(1, 3))

            # 最终保存 - 传统Excel格式
            excel_path = self.config.OUTPUT_DIR / self.config.EXCEL_FILENAME
            main_workbook.save(excel_path)
            logger.info(f"传统Excel文件已保存: {excel_path}")

            # 收集所有项目数据用于多格式输出
            all_projects_data = []
            for page_num in range(1, min(self.config.MAX_PAGES + 1, self.stats.pages_processed + 1)):
                current_page_url = self.config.get_page_url(page_num)
                page_html = self.askURL(current_page_url)
                if page_html:
                    projects_on_page, _ = self.parse_main_listing_page(page_html, 0)
                    all_projects_data.extend(projects_on_page)

            # 生成多格式输出
            if all_projects_data:
                stats_summary = self.stats.get_summary()
                saved_files = self.save_data_to_multiple_formats(
                    all_projects_data,
                    suffix="final",
                    stats=stats_summary
                )

                logger.info(f"\n🎉 多格式文件生成完成:")
                for format_type, file_path in saved_files.items():
                    logger.info(f"  {format_type.upper()}: {file_path}")

            # 输出统计信息
            self.print_stats()

            logger.info(f"\n--- 爬取完成 ---")
            logger.info(f"Total projects processed and attempted to save: {total_projects_processed}")
            logger.info(f"Final global excel row index reached: {excel_index_counter}")

            return True

        except Exception as e:
            logger.error(f"爬虫运行失败: {e}")
            self.stats.record_error("runtime_error", str(e))
            return False

    def run(self) -> bool:
        """运行爬虫（增强版接口）"""
        logger.info("开始运行摩点众筹爬虫 - 增强版接口")

        all_projects = []

        try:
            for page_num in range(1, self.config.MAX_PAGES + 1):
                logger.info(f"正在处理第 {page_num} 页...")

                page_url = self.config.get_page_url(page_num)
                html = self.make_request(page_url)

                if not html:
                    logger.warning(f"第 {page_num} 页获取失败，跳过")
                    continue

                # 简化版解析（用于增强版接口）
                projects = self.parse_listing_page_simple(html)
                all_projects.extend(projects)

                self.stats.pages_processed += 1

                logger.info(f"第 {page_num} 页处理完成，找到 {len(projects)} 个项目")

                # 定期保存
                if page_num % self.config.SAVE_INTERVAL == 0:
                    self.save_data_to_multiple_formats(all_projects, f"intermediate_page_{page_num}")

                # 添加延迟避免过于频繁的请求
                time.sleep(random.uniform(1, 3))

            # 最终保存 - 使用增强的多格式输出
            if all_projects:
                # 转换简化格式为完整格式（用于兼容）
                projects_data = []
                for i, project in enumerate(all_projects, 1):
                    project_row = [
                        i,  # 序号
                        project.get("link", ""),  # 项目link
                        project.get("id", ""),    # 项目6位id
                        project.get("title", ""), # 项目名称
                        project.get("image", "")  # 项目图
                    ]
                    # 填充其他字段为空值（简化版没有详细信息）
                    project_row.extend([""] * 29)  # 补充到34个字段
                    projects_data.append(project_row)

                stats_summary = self.stats.get_summary()
                saved_files = self.save_data_to_multiple_formats(
                    projects_data,
                    suffix="enhanced_final",
                    stats=stats_summary
                )

                logger.info(f"🎉 爬取完成！共处理 {len(all_projects)} 个项目")
                logger.info(f"📁 生成的文件:")
                for format_type, file_path in saved_files.items():
                    logger.info(f"  {format_type.upper()}: {file_path}")
            else:
                logger.warning("未获取到任何项目数据")

            # 输出统计信息
            self.print_stats()

            return True

        except Exception as e:
            logger.error(f"爬虫运行失败: {e}")
            self.stats.record_error("runtime_error", str(e))
            return False

    def parse_listing_page_simple(self, html: str) -> List[Dict[str, Any]]:
        """简化版列表页面解析（用于增强版接口）"""
        projects = []

        try:
            soup = BeautifulSoup(html, "html.parser")
            pro_field_div = soup.find('div', {'class': 'pro_field'})

            if not pro_field_div:
                logger.warning("未找到项目列表容器")
                return projects

            project_items = pro_field_div.find_all('li')
            logger.info(f"找到 {len(project_items)} 个项目")

            for item_li in project_items:
                project = self.parse_project_item_simple(item_li)
                if project:
                    projects.append(project)
                    self.stats.projects_found += 1

        except Exception as e:
            logger.error(f"解析列表页面失败: {e}")
            self.stats.record_error("parse_error", str(e))

        return projects

    def parse_project_item_simple(self, item_li) -> Optional[Dict[str, Any]]:
        """简化版项目项解析"""
        try:
            # 查找项目链接
            item_link = ""

            for a_tag in item_li.find_all('a'):
                href = a_tag.get('href', '')
                if '/item/' in href:
                    item_link = href
                    break

            if not item_link:
                return None

            if not item_link.startswith("http"):
                item_link = "https://zhongchou.modian.com" + item_link

            # 提取项目信息
            project_id = ""
            id_match = self.config.LINK_ID_PATTERN.search(item_link)
            if id_match:
                project_id = id_match.group(1)

            # 项目标题
            title = ""
            title_h3 = item_li.find('h3', class_='pro_title')
            if title_h3:
                title = title_h3.text.strip()

            # 项目图片
            img_src = ""
            img_tag = item_li.find('img')
            if img_tag and img_tag.get('src'):
                img_src = img_tag.get('src')

            # 跳过特定项目
            if "可汗游戏大会" in title:
                return None

            return {
                "link": item_link,
                "id": project_id,
                "title": title,
                "image": img_src
            }

        except Exception as e:
            logger.error(f"解析项目项失败: {e}")
            return None


def main(category: str = "all"):
    """主函数"""
    print("🚀 启动摩点爬虫管理系统 - 融合版")
    print(f"📂 爬取分类: {category}")

    # 创建配置
    config = ModianSpiderConfig(category)

    # 创建爬虫实例
    spider = ModianSpider(config)

    # 运行爬虫（使用完整版接口）
    success = spider.run_scraper()

    if success:
        print("✅ 爬虫运行完成")
    else:
        print("❌ 爬虫运行失败")

    return success


if __name__ == "__main__":
    main()
    print("爬取完毕！")
